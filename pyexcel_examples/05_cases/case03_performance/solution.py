"""
РЕШЕНИЕ: Оптимизация производительности на больших файлах

1. Используйте pandas.read_excel с параметром chunksize
2. Используйте openpyxl с использованием iter_rows()
3. Для очень больших файлов используйте специализированные библиотеки
"""

import pandas as pd
import pyexcel as pe
import openpyxl
import time

print("=" * 60)
print("РЕШЕНИЕ: Оптимизация производительности")
print("=" * 60)

# ✓ ВАРИАНТ 1: Использование pandas с chunksize
print("\n✓ ВАРИАНТ 1: pandas с потоковой обработкой (РЕКОМЕНДУЕТСЯ)")
start = time.time()

chunk_size = 1000
total_salary = 0
count = 0

for chunk in pd.read_excel('sample_large.xlsx', chunksize=chunk_size):
    if 'Salary' in chunk.columns:
        total_salary += chunk['Salary'].sum()
        count += len(chunk)

elapsed = time.time() - start
print(f"  Время обработки: {elapsed:.2f} сек")
print(f"  Обработано строк: {count}")
print(f"  Сумма зарплат: {total_salary:,.0f}")

# ✓ ВАРИАНТ 2: Использование openpyxl с iter_rows()
print("\n✓ ВАРИАНТ 2: openpyxl с потоковой обработкой")
start = time.time()

wb = openpyxl.load_workbook('sample_large.xlsx', data_only=True)
ws = wb.active

header = None
total_salary = 0
count = 0
salary_idx = -1

for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx == 0:
        header = row
        salary_idx = header.index('Salary') if 'Salary' in header else -1
    else:
        if salary_idx >= 0 and row[salary_idx]:
            try:
                total_salary += float(row[salary_idx])
                count += 1
            except (ValueError, TypeError):
                pass

elapsed = time.time() - start
print(f"  Время обработки: {elapsed:.2f} сек")
print(f"  Обработано строк: {count}")
print(f"  Сумма зарплат: {total_salary:,.0f}")

# ✓ ВАРИАНТ 3: Использование pandas с read_excel без загрузки всего в память
print("\n✓ ВАРИАНТ 3: pandas с одноразовым чтением (если нужны все данные)")
start = time.time()

# Читаем только нужные столбцы
df = pd.read_excel('sample_large.xlsx', usecols=['ID', 'Salary'])
total_salary = df['Salary'].sum()
count = len(df)

elapsed = time.time() - start
print(f"  Время обработки: {elapsed:.2f} сек")
print(f"  Обработано строк: {count}")
print(f"  Сумма зарплат: {total_salary:,.0f}")

# ✓ ВАРИАНТ 4: Стратегия выбора инструмента
print("\n✓ ВАРИАНТ 4: Выбор инструмента по размеру файла")
print("  < 1000 строк: pyexcel, pandas (любой способ)")
print("  1000-10000 строк: pandas с usecols (выбираем столбцы)")
print("  10000-100000 строк: pandas с chunksize (потоковая обработка)")
print("  > 100000 строк: openpyxl или специализированные библиотеки")

# ✓ ВАРИАНТ 5: Практические рекомендации
print("\n✓ ВАРИАНТ 5: Лучшие практики")
print("  1. Используйте usecols для чтения только нужных столбцов")
print("  2. Используйте chunksize для больших файлов")
print("  3. Используйте data_only=True в openpyxl")
print("  4. Преобразуйте в нужный формат один раз")
print("  5. Рассмотрите использование баз данных (sqlite, postgresql)")

# Пример: оптимальный код для работы с большим файлом
print("\n✓ ПРИМЕР: Оптимальный подход для аналитики")

def analyze_large_file(filepath, chunk_size=5000):
    """Анализирует большой файл экономно по памяти"""
    stats = {
        'total_rows': 0,
        'total_salary': 0,
        'avg_salary': 0,
        'departments': {},
    }

    for chunk in pd.read_excel(filepath, chunksize=chunk_size):
        stats['total_rows'] += len(chunk)

        if 'Salary' in chunk.columns:
            stats['total_salary'] += chunk['Salary'].sum()

        if 'Department' in chunk.columns:
            dept_counts = chunk['Department'].value_counts()
            for dept, count in dept_counts.items():
                stats['departments'][dept] = stats['departments'].get(dept, 0) + count

    if stats['total_rows'] > 0:
        stats['avg_salary'] = stats['total_salary'] / stats['total_rows']

    return stats

start = time.time()
result = analyze_large_file('sample_large.xlsx')
elapsed = time.time() - start

print(f"  Анализ завершён за {elapsed:.2f} сек")
print(f"  Всего строк: {result['total_rows']}")
print(f"  Средняя зарплата: {result['avg_salary']:,.0f}")
print(f"  Количество отделов: {len(result['departments'])}")

