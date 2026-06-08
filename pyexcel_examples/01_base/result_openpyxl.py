from openpyxl import load_workbook

wb = load_workbook("prices.xlsx")
sheet = wb.active
data = []

for row in sheet.iter_rows(values_only=True):
   data.append(list(row))
print(data)
